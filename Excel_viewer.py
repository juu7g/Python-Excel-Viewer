"""
エクセルビューア
"""

from datetime import datetime
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.font as tkFont
import csv, itertools, re, os
import openpyxl as oxl
import openpyxl.chartsheet as oxlch
import openpyxl.utils.datetime as oxldt
from tkinter import filedialog

class ListView(ttk.Frame):
    """
    Excelの結果をリストビューで表示する
    """
    def __init__(self, master):
        """
        画面の作成
        上のFrame: 入力用
        下のFrame: notebookで出力
        """
        super().__init__(master)
        self.excel_op = ExcelOp()
        self.csv_mng = CsvManage()
        self.u_frame = tk.Frame(bg="blue")      # 背景色を付けて配置を見る
        self.b_frame = tk.Frame(bg="green")     # 背景色を付けて配置を見る
        self.note = ttk.Notebook(self.b_frame)
        self.u_frame.pack(fill=tk.X)
        self.b_frame.pack(fill=tk.BOTH, expand=True)
        self.note.pack(fill=tk.BOTH, expand=True)
        self.create_input_frame(self.u_frame)

    def fixed_map(self, option):
        # Fix for setting text colour for Tkinter 8.6.9
        # From: https://core.tcl.tk/tk/info/509cafafae
        #
        # Returns the style map for 'option' with any styles starting with
        # ('!disabled', '!selected', ...) filtered out.

        # style.map() returns an empty list for missing options, so this
        # should be future-safe.
        return [elm for elm in self.style.map('Treeview', query_opt=option) if
            elm[:2] != ('!disabled', '!selected')]

    def create_input_frame(self, parent):
        """
        入力項目の画面の作成
        上段：入力ファイルパス、ファイル選択ボタン、開くボタン、日付変換チェックボックス、CSV出力ボタン
        下段：メッセージ
        """
        self.lbl_excel = tk.Label(parent, text="Excel:")
        self.var_excel_path = tk.StringVar(value="")
        self.ety_excel_path = tk.Entry(parent, textvariable=self.var_excel_path)
        self.btn_f_sel = tk.Button(parent, text="ファイル選択", command=self.select_file)
        self.btn_open = tk.Button(parent, text="開く", command=self.open_excel)
        self.var_dt = tk.IntVar(value=0)
        self.ckb_dt = tk.Checkbutton(parent, text="日付変換", variable=self.var_dt)
        self.btn_csv = tk.Button(parent, text="CSV出力", command=self.write_csv, state="disable")
        self.msg = tk.StringVar(value="msg")
        self.lbl_msg = tk.Label(parent
                                , textvariable=self.msg
                                , justify=tk.LEFT
                                , font=("Fixedsys", 11)
                                , relief=tk.RIDGE
                                , anchor=tk.W)
        self.lbl_msg.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)    # 先にpackしないと下に配置されない
        self.lbl_excel.pack(side=tk.LEFT, fill=tk.BOTH)
        self.ety_excel_path.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.btn_csv.pack(side=tk.RIGHT, fill=tk.Y)
        self.ckb_dt.pack(side=tk.RIGHT, fill=tk.Y)
        self.btn_open.pack(side=tk.RIGHT)
        self.btn_f_sel.pack(side=tk.RIGHT)
        self.ety_excel_path.bind("<Return>", self.open_excel)   #Enterキーを押しても動作するように

    def create_tree_frame(self, parent:ttk.Notebook, tab_name="") -> ttk.Treeview:
        """
        Treeviewとスクロールバーを持つframeを作成し、notebookにaddする。
        frameは、Treeviewとスクロールバーをセットする
        Treeviewは、listview形式、行は縞模様
        Args:
            ttk.Notebook:   ttk.Notebook
            string:         tab_name
        Returns:
            Treeview:       ツリービュー
        """
        # tagを有効にするためstyleを更新 tkinter8.6?以降必要みたい
        # 表の文字色、背景色の設定に必要
        self.style = ttk.Style()
        self.style.map('Treeview', foreground=self.fixed_map('foreground')
                                 , background=self.fixed_map('background'))
        # タブごとのスタイルの設定
        self.style.configure(tab_name + ".Treeview")
        # frameの作成。frameにTreeviewとScrollbarを配置する
        frame1 = tk.Frame(parent, bg="cyan")
        # Treeviewの作成
        treeview1 = ttk.Treeview(frame1, style=tab_name + ".Treeview")
        treeview1["show"] = "headings"      # 表のデータ列だけを表示する指定
        treeview1.tag_configure("odd", background="ivory2")     # 奇数行の背景色を指定するtagを作成
        # 水平スクロールバーの作成
        h_scrollbar = tk.Scrollbar(frame1, orient=tk.HORIZONTAL, command=treeview1.xview)
        treeview1.configure(xscrollcommand=h_scrollbar.set)
        # 垂直スクロールバーの作成
        v_scrollbar = tk.Scrollbar(frame1, orient=tk.VERTICAL, command=treeview1.yview)
        treeview1.configure(yscrollcommand=v_scrollbar.set)
        # pack expandがある方を後にpackしないと他が見えなくなる
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)          # 先にパックしないと表示されない
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)           # 先にパックしないと表示されない
        treeview1.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        parent.add(frame1, text=tab_name)
        return treeview1

    def update_tree_column(self, tree:ttk.Treeview, columns:list):
        """
        TreeViewの列定義と見出しを設定
        見出しの文字長で列幅を初期設定
        Args:
            Treeview:   treeviewオブジェクト
            list:       列名のリスト
        """
        tree["columns"] = columns                  # treeviewの列定義を設定
        font1 = tkFont.Font()
        for col_name in columns:
            tree.heading(col_name, text=col_name)  # 見出しの設定
            width1 = font1.measure(col_name) + 10  # 見出しの文字幅をピクセルで取得
            # width1 = min(width1, 200)              # 見出しの幅が200pxより大きい時は200pxにする
            tree.column(col_name, width=width1)    # 見出し幅の設定

    def update_tree_by_result(self, tree:ttk.Treeview, rows:list):
        """
        rows(エクセルのデータ)をTreeViewに設定
        要素の文字幅が見出しの文字幅より長い場合は、列幅を変更する。
        奇数列の背景色を変更
        Args:
            Treeview:   Treeviewインスタンス
            list:       Excel実行結果セット(行リストの列リスト)
        """
        if not rows:    # 要素が無ければ戻る
            return
        font1 = tkFont.Font()
        # 要素の長さにより列幅を修正
        for i, _ in enumerate(rows[0]):     # 列数分回す(1行目の要素数分)
            # 同じ列のデータをリストにし列の値の長さを求め、最大となる列のデータを求める。
            # 値は数字もあるので文字に変換し長さを求める。また、Noneは'None'となるので'    'とする。
            max_str = max([x[i] for x in rows], key=lambda x:len(str(x))) or "    "
            # 求めたものが文字列だったら、改行された状態での最大となるデータを求める。
            # 厳密にはこの状態で最大となるデータを探さなければならないが割愛
            if type(max_str) is str:
                max_str = max(max_str.split("\n"), key=len)
            width1 = font1.measure(max_str) + 10   # 文字幅をピクセルで取得
            header1 = tree.column(tree['columns'][i], width=None) # 現在の幅を取得
            # 設定済みの列幅より列データの幅の方が大きいなら列幅を再設定
            if width1 > header1:
                tree.column(tree['columns'][i], width=width1)    # 見出し幅の再設定
                # print(f"幅の再設定 幅:{width1}、値:{max_str}")   # debug用
        
        # treeviewに要素追加。背景はtagを切り替えて設定
        tree.delete(*tree.get_children())   # Treeviewをクリア
        for i, row in enumerate(rows):
            tags1 = []              # tag設定値の初期化
            if i & 1:               # 奇数か? i % 2 == 1:
                tags1.append("odd") # 奇数番目(treeviewは0始まりなので偶数行)だけ背景色を変える(oddタグを設定)
            tree.insert("", tk.END, values=row, tags=tags1)     # Treeviewに1行分のデータを設定

    def open_excel(self, event=None):
        """
        エクセルファイルを開き、notebookにシートの内容をタブとして追加する
        タブにはTreeviewを追加し、シートのデータを追加する
        データの幅でTreeviewの列の幅を設定する
        データの行数でTreeviewの行の高さを設定する(行ごとにはできないので一番高い行に合わせる)
        """
        excel_path = self.var_excel_path.get()
        self.dict_tables = self.excel_op.get_excel_workbook(excel_path, self.var_dt.get())
        self.msg.set(self.excel_op.msg)

        # notebookの既存のタブを削除
        while self.note.tabs():
            self.note.forget("current")

        for sheet_name1 in self.dict_tables:
            # noteにTreeviewを持つタブを追加する
            treeview1 = self.create_tree_frame(self.note, sheet_name1)

            # 見出しの文字長で列幅を初期設定、treeviewのカラム幅を文字長に合わせて調整
            self.update_tree_column(treeview1, self.dict_tables.get(sheet_name1)[1])

            # rowsをTreeViewに設定、要素の文字幅が見出しの文字幅より長い場合は、列幅を変更する。偶数列の背景色を変更
            self.update_tree_by_result(treeview1, self.dict_tables.get(sheet_name1)[0])

            # 一番行数の多い行に合わせて高さを設定する
            # ２次元のデータを平坦化しstr型だけを抽出する
            cells = [s for s in itertools.chain.from_iterable(self.dict_tables.get(sheet_name1)[0]) if type(s) is str]
            if not cells:
                continue    # 対象がない場合は抜ける
            # 抽出したリストの要素の中で改行の数の最も多い要素を取得
            longest_cell = max(cells, key=lambda x:x.count("\n"))
            max_row_lines = longest_cell.count("\n") + 1             # 改行の数を数える
            # Treeviewの行の高さを変更
            self.style.configure(sheet_name1 + ".Treeview", rowheight = 18 * max_row_lines)
        
        # 表示したらself.dict_tablesができるのでCSV出力可にする
        self.btn_csv.configure(state="active")


    def select_file(self, event=None):
        """
        ファイル選択ダイアログを表示。選択したファイルパスを保存
        """
        file_path = filedialog.askopenfilename(filetypes=[("XLSX/XLSM", ".xlsx .xlsm"), ("All", "*")])
        self.var_excel_path.set(file_path)

    def write_csv(self, event=None):
        """
        表示したデータをCSV出力
        エクセルのファイル名にシート名を付加したcsvファイルに出力
        """
        excel_path = self.var_excel_path.get()
        excel_path = os.path.splitext(excel_path)[0]    # 拡張子を除く
        for sheet_name in self.dict_tables:
            table = self.dict_tables.get(sheet_name)
            # csvへ出力 見出しは表示用のダミーなので出力しない
            self.csv_mng.write_csv(excel_path + "_" + sheet_name, None, table[0])
        if self.csv_mng.msg:
            self.msg.set(self.csv_mng.msg)  # エラーがあった場合、エラー内容を表示
        else:
            self.msg.set("csv出力完")


class ExcelOp():
    """
    Excelデータの操作を行う
    """
    def __init__(self):
        self.msg = ""   # メッセージ受渡し用

    def conv_format_excel2python(self, nf:str) -> str:
        """
        エクセル用の書式をPythonの書式に変換
        現在、日付のみ対応
        Args:
            str:    エクセルでの書式
        Returns:
            str:    Pythonでの書式
        """
        nf = re.sub(";.+", "", nf)     # ;以降を削除。エクセルは書式を「;」で区切って複数指定できるので２番目以降は無視
        nf = nf.replace('"', "")       # '"'を削除
        nf = nf.replace("yyyy", "%Y").replace("yy", "%y")   # 年
        nf = re.sub("[m]+", "%m", nf)  # 月。分との区別は未対応
        nf = re.sub("[d]+", "%d", nf)  # 日
        return nf
    
    def conv_cell_excel2python(self, cell, conv_dt:bool) -> str:
        """
        エクセルのデータで変換が必要なものを変換する
        - シリアル値が返っている日付データ(ただし、自動判別できないので指定してもらう)
        - 書式が設定されている日付
        Args:
            cell:   openpyxlのcell
            bool:   日付のシリアル値を変換するかどうか
        Returns:
            str:    セルの表示用の内容
        """
        # 日付のシリアル値を変換(エクセルでユーザー定義の書式が設定された場合)
        if conv_dt and cell.number_format == "General" and cell.data_type == "n":
            return oxldt.from_excel(cell.value)
        # 書式のある日付を変換
        if cell.data_type == "d" and cell.number_format != "General" and cell.number_format != None:
            _nf = self.conv_format_excel2python(cell.number_format)
            return cell.value.strftime(_nf)

    def get_excel_workbook(self, file_name:str, conv_dt:bool) -> dict:
        """
        エクセルファイルを読みデータを返す
        Args:
            str:    ファイル名
        Returns:
            dict:   エクセルのrowsとカラム定義をシート名をキーにした辞書
        """
        try:
            self.msg = ""   # メッセージクリア
            tables = {}
            wb = oxl.load_workbook(filename=file_name, read_only=True, data_only=True)
            sheet_names = wb.sheetnames    # シート名のリストを取得
            for ws_name in sheet_names:   # シート名で回す
                ws = wb[ws_name]          # シート名でシートの取得
                # Chartsheetの場合、コメントだけにする
                if isinstance(ws, oxlch.Chartsheet):
                    tables[ws_name] = ([("Chartsheet",)], [1])
                    print("chartsheet")     # debug用
                    continue
                # 寸法が正しくない場合、再計算する
                if ws.calculate_dimension(True) == "A1:A1":
                    # 空のシートはcalculate_dimensionでエラーになるので除く
                    if ws["A1"].value is not None:
                        ws.reset_dimensions()
                        ws.calculate_dimension(True)
                # 列定義の作成(現状は1開始の整数)
                columns1 = [i for i in range(1, ws.max_column + 1)]    # 列定義を列数分行う。1スタート
                # セル値の取得  DateTimeのシリアル値の変換も
                rows1 = list(list(x) for x in ws.values)    # 全ての値を取得 変更できるようにリストに変換
                for r, row in enumerate(ws.rows):
                    for c, cell1 in enumerate(row):
                        conv_value = self.conv_cell_excel2python(cell1, conv_dt)
                        # 変換されていたら置き換える
                        if conv_value:
                            rows1[r][c] = conv_value
                tables[ws_name] = (rows1, columns1)
            wb.close()
        except Exception as e:
            self.msg = e
        finally:
            return tables
    
class CsvManage():
    """
    csv操作管理用
    """
    def __init__(self):
        self.msg = ""   # メッセージ受渡し用

    def write_csv(self, path, header:list, rows:list):
        """
        csvファイルにrowsを出力
        ファイル名は「xlsxファイル名_シート名.csv」
        """
        # self.msg = ""
        try:
            with open(path + ".csv", encoding="cp932", mode="w", newline="") as f:   # cp932 or utf_8-sig
                writer_ = csv.writer(f)
                if header:
                    writer_.writerow(header)
                writer_.writerows(rows)
        except Exception as e:
            self.msg = "CSVエラー" + e

if __name__ == '__main__':
    root = tk.Tk()              # トップレベルウィンドウの作成
    root.title("Excel viewer")  # タイトル
    root.geometry("600x600")    # サイズ
    listview = ListView(root)   # ListViewクラスのインスタンス作成
    root.mainloop()
